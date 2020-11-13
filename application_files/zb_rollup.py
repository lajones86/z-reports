import zc_ledger as Ledger
import zj_investments as Investments
import zm_ledger as ReadLedger
import zr_excel as Excel
import zr_io as Io
import zl_liabilities as Liabilities
import zh_portfolio_beta as PortfolioBeta

import os
import openpyxl
import xlsxwriter

def write_overview_tab(xl_workbook, xl_worksheet, ledger_tabs, investments_tab, overview_entries):

    custom_cells = overview_entries[0]
    assets = overview_entries[1]
    liabilities = overview_entries[2]

    amount_format = xl_workbook.add_format({"num_format" : "#,##0.00" ";(#,##0.00)"})
    xl_section = xl_workbook.add_format({"bold" : True, "underline" : True, "center_across" : True})
    xl_header = xl_workbook.add_format({"italic" : True, "underline" : True, "center_across" : True})
    date_format = xl_workbook.add_format({"num_format" : "YYYY-MM-DD"})
    italic = xl_workbook.add_format({"italic" : True})
    bold = xl_workbook.add_format({"bold" : True})
    bold_italic = xl_workbook.add_format({"bold" : True, "italic" : True})
    bold_italic_underline = xl_workbook.add_format({"bold" : True, "italic" : True, "underline" : True})

    checking_formula = "=0"
    savings_formula = "=0"
    credit_card_formula = "=0"

    checking_adj_formula = "=0"
    savings_adj_formula = "=0"
    credit_card_adj_formula = "=0"

    for tab in ledger_tabs:
        if Ledger.get_tab_type(tab) == "liability":
            credit_card_formula += "+'%s'!%s" % (tab.name, tab.cc_total_cell)
            credit_card_adj_formula += "+'%s'!%s" % (tab.name, tab.cc_total_adj_cell)
        elif Ledger.get_tab_type(tab) == "asset":
            for account in tab.checking:
                checking_formula += "+'%s'!%s" % (tab.name, account.balance_cell)
                checking_adj_formula += "+'%s'!%s" % (tab.name, account.balance_adj_cell)
            for account in tab.savings:
                savings_formula += "+'%s'!%s" % (tab.name, account.balance_cell)
                savings_adj_formula += "+'%s'!%s" % (tab.name, account.balance_adj_cell)

    xl_worksheet.merge_range("A1:D1", "Summary", xl_section)
    xl_worksheet.write(1, 0, "Item", xl_header)
    xl_worksheet.write(1, 1, "Balance/Equity", xl_header)
    xl_worksheet.write(1, 2, "Pending Adj", xl_header)
    xl_worksheet.write(1, 3, "Adj Balance/Equity", xl_header)
    xl_worksheet.write(2, 0, "Checking Total", italic)
    xl_worksheet.write_formula(2, 1, checking_formula)
    xl_worksheet.write_formula(2, 2, checking_adj_formula)
    xl_worksheet.write_formula(2, 3, "=B3+C3")
    xl_worksheet.write(3, 0, "Savings Total", italic)
    xl_worksheet.write_formula(3, 1, savings_formula)
    xl_worksheet.write_formula(3, 2, savings_adj_formula)
    xl_worksheet.write_formula(3, 3, "=B4+C4")
    xl_worksheet.write(4, 0, "Credit Card Total", italic)
    xl_worksheet.write_formula(4, 1, credit_card_formula)
    xl_worksheet.write_formula(4, 2, credit_card_adj_formula)
    xl_worksheet.write_formula(4, 3, "=B5+C5")

    xl_worksheet.write(5, 0, "Cash Flows Total", bold)
    xl_worksheet.write_formula(5, 1, "=B3+B4-B5")
    xl_worksheet.write_formula(5, 2, "=C3+C4-C5")
    xl_worksheet.write_formula(5, 3, "=B6+C6")

    xl_worksheet.write(7, 0, "Cash", italic)
    cash_entry = custom_cells[0]
    xl_worksheet.write(7, 1, round(cash_entry[1], 2))
    xl_worksheet.write_formula(7, 2, "0")
    xl_worksheet.write_formula(7, 3, "=B8+C8")

    xl_worksheet.write(8, 0, "Total Liquid", bold)
    xl_worksheet.write_formula(8, 1, "=B6+B8")
    xl_worksheet.write_formula(8, 2, "=C6+C8")
    xl_worksheet.write_formula(8, 3, "=D6+D8")

    xl_worksheet.write(10, 0, "Brokerage Accounts", italic)
    xl_worksheet.write_formula(10, 1, "='Investments'!%s" % investments_tab.broker_total_cell)
    xl_worksheet.write_formula(10, 2, "0")
    xl_worksheet.write_formula(10, 3, "=B11+C11")

    xl_worksheet.write(11, 0, "Cryptocurrency", italic)
    xl_worksheet.write_formula(11, 1, "='Investments'!%s" % investments_tab.crypto_total_cell)
    xl_worksheet.write_formula(11, 2, "0")
    xl_worksheet.write_formula(11, 3, "=B12+C12")

    xl_worksheet.write(12, 0, "Metals", italic)
    xl_worksheet.write_formula(12, 1, "='Investments'!%s" % investments_tab.metal_total_cell)
    xl_worksheet.write_formula(12, 2, "0")
    xl_worksheet.write_formula(12, 3, "=B13+C13")

    xl_worksheet.write(13, 0, "Bonds", italic)
    xl_worksheet.write_formula(13, 1, "='Investments'!%s" % investments_tab.bond_total_cell)
    xl_worksheet.write_formula(13, 2, "0")
    xl_worksheet.write_formula(13, 3, "=B14+C14")

    xl_worksheet.write(14, 0, "Total Holdings", bold)
    xl_worksheet.write_formula(14, 1, "=SUM(B11:B14)")
    xl_worksheet.write_formula(14, 2, "=SUM(C11:C14)")
    xl_worksheet.write_formula(14, 3, "=SUM(D11:D14)")

    xl_worksheet.write(15, 0, "Assets Plus Cash Flows", bold_italic)
    xl_worksheet.write_formula(15, 1, "=B15+B9")
    xl_worksheet.write_formula(15, 2, "=C15+C9")
    xl_worksheet.write_formula(15, 3, "=D15+D9")

    #do liability/mortgage stuff here
    xl_row = 17
    extracted_liabilities = Liabilities.get_liabilities()
    for summary in extracted_liabilities.summaries_list:
        xl_worksheet.write(xl_row, 0, summary.name, italic)
        xl_worksheet.write(xl_row, 1, summary.balance)
        xl_worksheet.write_formula(xl_row, 2, "=0")
        xl_worksheet.write_formula(xl_row, 3, "=B%s+C%s" % (str(xl_row + 1), str(xl_row + 1)))
        xl_row += 1
    xl_worksheet.write(xl_row, 0, "Total Liabilities", bold)
    xl_worksheet.write_formula(xl_row, 1, "=SUM(B18:B%s)" % str(xl_row))
    xl_worksheet.write_formula(xl_row, 2, "=SUM(C18:C%s)" % str(xl_row))
    xl_worksheet.write_formula(xl_row, 3, "=B%s+C%s" % (str(xl_row + 1), str(xl_row + 1)))
    xl_row += 1
    xl_worksheet.write(xl_row, 0, "Assets + CF Less Liabilities", bold_italic)
    xl_worksheet.write_formula(xl_row, 1, "=B16-B%s" % str(xl_row))
    xl_worksheet.write_formula(xl_row, 2, "=C16-C%s" % str(xl_row))
    xl_worksheet.write_formula(xl_row, 3, "=B%s+C%s" % (str(xl_row + 1), str(xl_row + 1)))


    xl_row += 2
    xl_worksheet.write(xl_row, 0, "Other Assets", bold)
    xl_worksheet.write_formula(xl_row, 1, "=SUM(G:G)")
    xl_worksheet.write(xl_row, 2, 0)
    xl_worksheet.write_formula(xl_row, 3, "=B%s+C%s" % (str(xl_row + 1), str(xl_row + 1)))

    xl_row += 1
    xl_worksheet.write(xl_row, 0, "Other Liabilities", bold)
    xl_worksheet.write_formula(xl_row, 1, "=SUM(J:J)")
    xl_worksheet.write(xl_row, 2, 0)
    xl_worksheet.write_formula(xl_row, 3, "=B%s+C%s" % (str(xl_row + 1), str(xl_row + 1)))

    xl_row += 1
    xl_worksheet.write(xl_row, 0, "Other Total", bold_italic)
    xl_worksheet.write_formula(xl_row, 1, "=B%s-B%s" % (str(xl_row - 1), str(xl_row)))
    xl_worksheet.write_formula(xl_row, 2, "=C%s-C%s" % (str(xl_row - 1), str(xl_row)))
    xl_worksheet.write_formula(xl_row, 3, "=B%s+C%s" % (str(xl_row + 1), str(xl_row + 1)))

    xl_row += 2
    xl_worksheet.write(xl_row, 0, "Total Est Net Worth", bold_italic_underline)
    xl_worksheet.write_formula(xl_row, 1, "=B%s+B%s" % (str(xl_row - 1), str(xl_row - 5)))
    xl_worksheet.write_formula(xl_row, 2, "=C%s+C%s" % (str(xl_row - 1), str(xl_row - 5)))
    xl_worksheet.write_formula(xl_row, 3, "=B%s+C%s" % (str(xl_row + 1), str(xl_row + 1)))


    xl_worksheet.merge_range("F1:G1", "Other Assets", xl_section)
    xl_worksheet.write(1, 5, "Description", xl_header)
    xl_worksheet.write(1, 6, "Amount", xl_header)
    xl_row = 1
    for i in assets:
        xl_row += 1
        xl_worksheet.write(xl_row, 5, i[0])
        xl_worksheet.write(xl_row, 6, i[1])

    xl_worksheet.merge_range("I1:J1", "Other Liabilities", xl_section)
    xl_worksheet.write(1, 8, "Description", xl_header)
    xl_worksheet.write(1, 9, "Amount", xl_header)
    xl_row = 1
    for i in liabilities:
        xl_row += 1
        xl_worksheet.write(xl_row, 8, i[0])
        xl_worksheet.write(xl_row, 9, i[1])

    xl_worksheet.set_column(0, 0, 22)
    xl_worksheet.set_column(1, 1, 14, amount_format)
    xl_worksheet.set_column(2, 2, 12, amount_format)
    xl_worksheet.set_column(3, 3, 18, amount_format)
    xl_worksheet.set_column(4, 4, 5)
    xl_worksheet.set_column(5, 5, 20)
    xl_worksheet.set_column(6, 6, 18, amount_format)
    xl_worksheet.set_column(7, 7, 5)
    xl_worksheet.set_column(8, 8, 20)
    xl_worksheet.set_column(9, 9, 18, amount_format)

    return(None)


def get_custom_cells():
    custom_cells = []
    assets = []
    liabilities = []

    #copied from zm_ledger.load_ledger_entries()
    #lazy, lazy
    current_ledger = ReadLedger.get_ledger_path()
    prior_ledger = ReadLedger.get_ledger_path(prior = True)

    if os.path.isfile(current_ledger) and os.path.isfile(prior_ledger):
        Io.error("Current and prior ledgers both exist. Possibly unarchived prior ledger. Aborting.")    
    if not os.path.isfile(current_ledger) and not os.path.isfile(prior_ledger):
        print("No existing ledgers found. Writing empty ledger.")     
        return([["Cash", round(float(0), 2)]])

    if os.path.isfile(current_ledger):
        print("Loading current month ledger overview.")
        ledger = current_ledger

    if os.path.isfile(prior_ledger):
        print("Loading prior month ledger overview.")
        ledger = prior_ledger

    xl_workbook = openpyxl.load_workbook(ledger)
    try:
        xl_worksheet = xl_workbook["Overview"]
    except KeyError:
        print("No overview sheet found.")
        return([["Cash", round(float(0), 2)]])
    cash_value = (xl_worksheet.cell(row = 8, column = 2)).value
    try:
        cash_value = round(float(cash_value), 2)
    except (TypeError, ValueError):
        cash_value = round(float(0), 2)
    custom_cells.append(["Cash", cash_value])

    for row in range(3, xl_worksheet.max_row + 1):
        blanks = [None, ""]
        asset_desc = (xl_worksheet.cell(row = row, column = 6)).value
        asset_amt = (xl_worksheet.cell(row = row, column = 7)).value
        liab_desc = (xl_worksheet.cell(row = row, column = 9)).value
        liab_amt = (xl_worksheet.cell(row = row, column = 10)).value
        if not asset_desc in blanks or not asset_amt in blanks:
            assets.append([asset_desc, round(float(asset_amt), 2)])
        if not liab_desc in blanks or not liab_amt in blanks:
            liabilities.append([liab_desc, round(float(liab_amt), 2)])        

    return([custom_cells, assets, liabilities])

def main():

    overview_entries = get_custom_cells()

    xl_workbook = xlsxwriter.Workbook(ReadLedger.get_ledger_path())

    overview_tab = xl_workbook.add_worksheet("Overview")

    ledger_account_collection = Ledger.get_accounts()
    Ledger.archive_ledger()
    ledger_tabs = Ledger.write_ledger(ledger_account_collection, xl_workbook)

    investments = Investments.get_investments()
    investments_tab = Investments.write_worksheet(xl_workbook, investments)

    write_overview_tab(xl_workbook, overview_tab, ledger_tabs, investments_tab, overview_entries)

    Excel.save_workbook(xl_workbook, ReadLedger.get_ledger_path())

    PortfolioBeta.main(investments = investments)


if __name__ == "__main__":
    main()
