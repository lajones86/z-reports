import zr_config as Config
import zr_io as Io
import zr_financial_concepts as Concepts

import openpyxl
from datetime import datetime
from dateutil.relativedelta import relativedelta
import os

def get_ledger_path(prior = None):
    ledger_dir = Config.get_path("money_dir")
    base_time = datetime.today()
    if prior == True:
        base_time = base_time - relativedelta(months = 1)
    file_date = base_time.strftime("%Y-%m")
    ledger_path = os.path.join(ledger_dir, (file_date + "-Ledger.xlsx"))
    return(ledger_path)


def get_entries(xl_workbook, sheet_name, section_id):
    print("Loading %s, %s" % (sheet_name, section_id))
    entries = []
    xl_worksheet = xl_workbook[sheet_name]
    base_col = None
    for i in range(1, xl_worksheet.max_column + 1):
        cell = xl_worksheet.cell(row = 1, column = i)
        if cell.value == section_id:
            base_col = i
            break

    if base_col == None:
        return([])

    for row in range(3, xl_worksheet.max_row + 1):
        date = (xl_worksheet.cell(row = row, column = base_col)).value
        description = (xl_worksheet.cell(row = row, column = base_col + 1)).value
        amount = (xl_worksheet.cell(row = row, column = base_col + 2)).value
        if date != None or description != None or amount != None:
            entries.append(Concepts.LedgerEntry(date, description, amount))

    return(entries)


def load_rec_info(xl_workbook, account, sheet_name):
    xl_worksheet = xl_workbook[sheet_name]
    account.rec_date = (xl_worksheet.cell(row = 5, column = 1)).value
    account.bank_balance = (xl_worksheet.cell(row = 5, column = 3)).value


def load_ledger_entries(account):
    current_ledger = get_ledger_path()
    prior_ledger = get_ledger_path(prior = True)

    if os.path.isfile(current_ledger) and os.path.isfile(prior_ledger):
        Io.error("Current and prior ledgers both exist. Possibly unarchived prior ledger. Aborting.")    
    if not os.path.isfile(current_ledger) and not os.path.isfile(prior_ledger):
        print("No existing ledgers found. Writing empty ledger.")
        return(account)

    roll_forward = False

    if os.path.isfile(current_ledger):
        print("Loading current month ledger for %s." % account.name)
        ledger = current_ledger

    if os.path.isfile(prior_ledger):
        print("Loading prior month ledger for %s." % account.name)
        ledger = prior_ledger
        roll_forward = True
    
    xl_workbook = openpyxl.load_workbook(ledger)

    #get book entries
    if account.account_type == "Credit Card":
        book_tab = "Credit Cards"
    else:
        book_tab = account.institution
    rec_tab = account.name + " - Rec"

    load_rec_info(xl_workbook, account, rec_tab)

    book_entries = get_entries(xl_workbook, book_tab, account.name)
    book_adjustments = get_entries(xl_workbook, rec_tab, "Book Adjustments")
    bank_adjustments = get_entries(xl_workbook, rec_tab, "Bank Adjustments")

    if roll_forward == False:
        balance_entry = book_entries[0]
        account.book_balance = float(balance_entry.amount)
        book_entries.remove(balance_entry)
        account.book_adjustments = book_adjustments

    else:
        balance = float(0)
        for entry in book_entries:
            balance += entry.amount
        account.book_balance = balance
        book_entries = []
        for adjustment in book_adjustments:
            if adjustment.date.year == datetime.now().year and adjustment.date.month == datetime.now().month:
                book_entries.append(adjustment)
        for entry in book_entries:
            book_adjustments.remove(entry)

    account.book_entries = book_entries
    account.book_adjustments = book_adjustments
    account.bank_adjustments = bank_adjustments

    if roll_forward == True:
        if account.rec_date.date() != datetime.now().date():
            print(account.rec_date.date())
            print(datetime.now().date())
            Io.error("Bad reconciliation date. Reconciliation must be done on day of roll forward.")

        book_rec_bal = float(account.book_balance)
        bank_rec_bal = float(account.bank_balance)

        for entry in account.book_entries:
            book_rec_bal += float(entry.amount)
        for entry in account.book_adjustments:
            book_rec_bal += float(entry.amount)
        for entry in account.bank_adjustments:
            bank_rec_bal += float(entry.amount)

        book_rec_bal = round(book_rec_bal, 2)
        bank_rec_bal = round(bank_rec_bal, 2)

        if bank_rec_bal != book_rec_bal:
            Io.error("Reconciliation failed. Accounts must be reconciled before rolling forward.")
